# -*- coding: utf-8 -*-
"""
Created on Sat Dec  5 11:57:16 2020

@author: pierr

"""

# =============================================================================
# Code permettant 
# =============================================================================

from openpyxl import load_workbook # module pour lire un doc excel
from openpyxl.utils import get_column_letter,column_index_from_string # pour convertir les colonnes en lettre


 
document = load_workbook("note_freq.xlsx") # on importe le doc a lire 
feuille = document.worksheets[0] # selectionne la première feuille, ici la seule
doc = document.active # permet la modification du document


# la, on indique le nombre de lignes et de colonnes du tableau
nb_row = feuille.max_row # on choppe le nombre de lignes
nb_column = feuille.max_column # on choppe le nombre de colonnes

# maintenant, on dois chopper toutes les infos de chaque ligne une par une 
note = [] # on créer un tableau dans lequel il y aura toutes les notes d'un clavier 11 octaves
for i in range(3,nb_row+1) : # on parcoure toutes les lignes sauf la première qui sont les titres
    i = nb_row+3 - i # on veut commencer par la première touche du piano 
    tempo = [] # tableau temporaire pour stocker toute la ligne à l'intérieur et ensuite l'envoyer dans contenu
    for l in range(1,nb_column+1) : # on parcoure toutes les colonnes de chacunes des lignes
        cellule = feuille[str(get_column_letter(l))+str(i)].value # on attribue la valeur de la cellule a cette variable
        tempo.append(cellule) # la, on ajoute au tableau temporaire le contenu des cellules
    note.append(tempo) # on l'ajoute dans la bonne liste 
print(note)
